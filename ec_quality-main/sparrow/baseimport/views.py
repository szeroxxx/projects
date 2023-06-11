# import ast
import csv
import logging
import os

import openpyxl
from base.models import AppResponse
from base.util import Util
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render
from messytables import CSVTableSet, XLSTableSet


def load_import_template(request, model_name):
    return render(request, "baseimport/import.html", {"model_name": model_name})


def generate_file_data(request):
    try:
        available_cross_ref_fields = None
        if request.method == "POST":
            model_name = request.POST["model"].strip()
            fields = []
            cross_refs = None
            dropdown_fields = None
            if model_name in Util.import_dropdown_field:
                dropdown_fields = Util.import_dropdown_field[model_name]
            mapping_columns = Util.import_model_fields[model_name]
            if model_name == "product":
                forms = ProductGroup_Form.objects.filter()
                for form in forms:
                    form_fields = Template_Form_Field.objects.filter(template_form_id=form.template_form_id)
                    for form_field in form_fields:
                        full_name = form.name.lower().replace(" ", "_") + "_" + form_field.name.lower().replace(" ", "_")
                        fields.append(full_name)

            if model_name == "contact":
                custom_field_headers = []
                custom_fields = ContactCustomField.objects.filter(name__isnull=False).values_list("name", flat=True).distinct()
                custom_field_headers = ["custom_field_" + str(x) for x in custom_fields] if len(custom_fields) > 0 else custom_field_headers
                fields += custom_field_headers

                if request.POST.get("is_subscriber_import") == "true":
                    new_dropdown_fields = [x for x in dropdown_fields if x["col_name"] == "email"]
                    dropdown_fields = new_dropdown_fields
                    new_mapping_columns = []
                    for mapping_column in mapping_columns:
                        if mapping_column["col_name"] == "email":
                            mapping_column["is_required"] = "true"
                        new_mapping_columns.append(mapping_column)
                    mapping_columns = new_mapping_columns

            if model_name == "alternatives":
                available_cross_ref_fields = CrossReference.objects.filter(is_deleted=False).values("id", "name")
                cross_refs = []
                for cross_ref in available_cross_ref_fields:
                    cross_refs.append(cross_ref["name"])

            if request.FILES["file"] is not None and ".csv" in str(request.FILES["file"]):
                fh = request.FILES["file"]
                delimiter = None
                if request.POST.get("delimiter", False):
                    delimiter = request.POST["delimiter"]
                    delimiter = str(delimiter)

                table_set = CSVTableSet(fh, delimiter)
                if delimiter is None:
                    dialect = csv.Sniffer().sniff(fh.read(1024).decode("utf-8"))
                    delimiter = dialect.delimiter
                row_set = table_set.tables[0]
                data = []
                for row in row_set:
                    row_data = []
                    is_row_null = True
                    for cell in row:
                        value = cell.value
                        if str(cell.type).lower() == "string":
                            value = cell.value.strip()
                        if str(cell.type).lower() == "datetime":
                            value = str(cell.value).strip()
                        row_data.append(value)
                        if is_row_null and value is not None and value != "":
                            is_row_null = False
                    if not is_row_null:
                        data.append(row_data)
                return HttpResponse(
                    AppResponse.get(
                        {
                            "code": 1,
                            "data": data,
                            "mapping_columns": mapping_columns,
                            "delimiter": delimiter,
                            "file_type": "csv",
                            "fields": fields,
                            "dropdown_fields": dropdown_fields,
                            "model_name": model_name,
                            "available_cross_ref_fields": cross_refs,
                        }
                    ),
                    content_type="json",
                )

            if request.FILES["file"] is not None and (".xls" in str(request.FILES["file"])):
                sheet_number = 0
                fh = request.FILES["file"]
                data = []
                mapping_columns.sort(key=lambda obj: obj["is_required"], reverse=True)
                if ".xlsx" not in str(request.FILES["file"]):
                    xlsclass = XLSTableSet
                    kwargs = {"encoding": "utf-8"}
                    file_type = "xls"

                    table_set = xlsclass(request.FILES["file"], **kwargs)
                    try:
                        row_set = table_set.tables[sheet_number]
                    except IndexError:
                        raise Exception("This file does not have sheet number %d" % (sheet_number + 1))

                    for row in row_set:
                        row_data = []
                        is_row_null = True
                        for cell in row:
                            value = cell.value
                            if isinstance(cell.value, str):
                                value = cell.value.strip()
                            if "date" in str(cell.type).lower():
                                value = str(cell.value).strip()
                                if value != "":
                                    value = str(cell.value.date())
                            row_data.append(value)
                            if is_row_null and value is not None and value != "":
                                is_row_null = False
                        if not is_row_null:
                            data.append(row_data)

                if ".xlsx" in str(request.FILES["file"]):
                    file_type = "xlsx"
                    book = openpyxl.load_workbook(request.FILES["file"], data_only=True)
                    sheet = book.worksheets[0]

                    # To count correct number of rows and columns from xlsx file with data inside them
                    row_count = 0
                    column_count = 0
                    row_count = len([row for row in sheet if not all([cell.value is None for cell in row])])
                    # Find columns count
                    for row in sheet:
                        row_info = []
                        for cell in row:
                            row_info.append(cell.value)
                        if len(list(filter(None, row_info))) > 0:
                            while row_info[-1] is None:
                                row_info.pop(-1)
                            if column_count < len(row_info):
                                column_count = len(row_info)

                    for row in range(1, row_count + 1):
                        row_data = []
                        is_row_null = True
                        for col in range(1, column_count + 1):
                            value = sheet.cell(row=row, column=col).value
                            if value is not None:
                                if sheet.cell(row=row, column=col).number_format == "0.00%":
                                    value = value * 100
                                if sheet.cell(row=row, column=col).is_date is True:
                                    value = str((sheet.cell(row=row, column=col).value).date())
                                row_data.append(str(value))
                            else:
                                row_data.append("")
                            if is_row_null and value is not None and value != "":
                                is_row_null = False
                        if not is_row_null:
                            data.append(row_data)
                if len(data) > 5001:
                    return HttpResponse(AppResponse.msg(0, "Maximum 5000 records are allowed in import."))
                return HttpResponse(
                    AppResponse.get(
                        {
                            "code": 1,
                            "data": data,
                            "mapping_columns": mapping_columns,
                            "delimiter": None,
                            "file_type": file_type,
                            "fields": fields,
                            "dropdown_fields": dropdown_fields,
                            "model_name": model_name,
                            "available_cross_ref_fields": cross_refs,
                        }
                    ),
                    content_type="json",
                )
    except Exception as e:
        manager.create_from_exception(e)
        logging.exception("Something went wrong.")
        return HttpResponse(AppResponse.msg(0, str(e)), content_type="json")


def export_sample_product(request, group_id):
    try:
        headers = []
        row_number = 1
        group_name = ""
        group = ProductGroup_Form.objects.filter(product_group_id=group_id).first()
        if group:
            group_name = group.name
            product_attrs = Template_Form_Field.objects.filter(template_form=group.template_form)
            for product_attr in product_attrs:
                name = (product_attr.name).replace(" ", "_")
                attribute = "attribute_" + group.name + "_" + name
                headers.append((attribute).lower())

        file_path = os.path.join(settings.BASE_DIR, "base", "static", "base", "sampletemplates", "products_import.xlsx")
        wb1 = openpyxl.load_workbook(file_path)
        ws1 = wb1.worksheets[0]

        row_count = ws1.max_row
        column_count = ws1.max_column

        for i in range(1, row_count):
            ws1.cell(row=row_number + 1, column=(column_count)).value = group_name
            row_number = row_number + 1

        row_number = 1
        for header in headers:
            ws1.cell(row=row_number, column=(column_count + 1)).value = header
            column_count = column_count + 1

        response = HttpResponse(content_type="application/ms-excel")
        fname = "sample_product.xlsx"
        response["Content-Disposition"] = "attachment; filename=%s" % fname
        wb1.save(response)
        return response

    except Exception as e:
        manager.create_from_exception(e)
        logging.exception("Something went wrong.")
        return HttpResponse(AppResponse.msg(0, str(e)), content_type="json")
