from django.shortcuts import render
from django.apps import apps
from django.db.models import Q
from django.http import HttpResponse, response, StreamingHttpResponse
from base.models import AppResponse
from base import views as base_views
from django.conf import settings
from django.core.files import File
from wsgiref.util import FileWrapper
from messytables import CSVTableSet, type_guess, \
  types_processor, headers_guess, headers_processor, \
  offset_processor, any_tableset,XLSTableSet,XLSXTableSet
import json
from base.util import Util
import csv
import ast
from io import StringIO
import xlwt
import openpyxl
from exception_log import manager
import logging
from io import BytesIO

def load_import_template(request, model_name):
    return render(request, 'baseimport/import.html',{'model_name':model_name})

def generate_file_data(request):
    try:
        if request.method == 'POST':
            model_name = request.POST['model'].strip()
            fields = []
            dropdown_fields = None
            if model_name in Util.import_dropdown_field:
                dropdown_fields  = Util.import_dropdown_field[model_name]
            mapping_columns = Util.import_model_fields[model_name]
            if model_name == 'product':
                forms = ProductGroup_Form.objects.filter()
                for form in forms:
                    form_fields = Template_Form_Field.objects.filter(template_form_id = form.template_form_id)
                    for form_field in form_fields:
                        full_name = form.name.lower().replace(" ", "_") + "_" + form_field.name.lower().replace(" ", "_")
                        fields.append(full_name)
                        
            if model_name == 'contact':
                custom_field_headers = []
                custom_fields = ContactCustomField.objects.filter(name__isnull=False).values_list('name', flat = True).distinct()
                custom_field_headers = ['custom_field_'+str(x) for x in custom_fields] if len(custom_fields) > 0 else custom_field_headers
                fields += custom_field_headers

                if request.POST.get('is_subscriber_import') == 'true':
                    new_dropdown_fields = [x for x in dropdown_fields if x['col_name'] == 'email']
                    dropdown_fields = new_dropdown_fields
                    new_mapping_columns = []
                    for mapping_column in mapping_columns:
                        if mapping_column['col_name'] == 'email':
                            mapping_column['is_required'] = "true"
                        new_mapping_columns.append(mapping_column)
                    mapping_columns = new_mapping_columns

            if request.FILES['file'] != None and '.csv' in str(request.FILES['file']):
                fh = request.FILES['file']
                c_ip = base_views.get_client_ip(request)
                delimiter = None
                if request.POST.get('delimiter',False):
                    delimiter = request.POST['delimiter']
                    delimiter = str(delimiter)

                table_set = CSVTableSet(fh,delimiter)
                if delimiter == None:
                    dialect = csv.Sniffer().sniff(fh.read(1024).decode("utf-8"))
                    delimiter = dialect.delimiter
                row_set = table_set.tables[0]
                data = []
                for row in row_set:
                    row_data = []
                    is_row_null = True
                    for cell in row:
                        value = cell.value
                        if str(cell.type).lower() == 'string':
                            value = cell.value.strip()
                        if str(cell.type).lower() == 'datetime':
                            value = str(cell.value).strip()
                        row_data.append(value)
                        if is_row_null and value != None and value !='':
                            is_row_null = False
                    if not is_row_null:
                        data.append(row_data)
                return HttpResponse(AppResponse.get({'code' : 1, 'data': data, 'mapping_columns': mapping_columns,'delimiter':delimiter,'file_type':'csv', 'fields': fields,'dropdown_fields':dropdown_fields, 'model_name':model_name }), content_type='json')  
            
            if request.FILES['file'] != None and ('.xls' in str(request.FILES['file'])):
                sheet_number = 0
                fh = request.FILES['file']
                data = []
                mapping_columns.sort(key = lambda obj: obj['is_required'], reverse=True)
                if '.xlsx' not in str(request.FILES['file']):
                    xlsclass = XLSTableSet
                    kwargs = { 'encoding': 'utf-8' }
                    file_type = 'xls'

                    table_set = xlsclass(request.FILES['file'], **kwargs)
                    try:
                        row_set = table_set.tables[sheet_number]
                    except IndexError:
                        raise Exception('This file does not have sheet number %d' %(sheet_number + 1))

                    for row in row_set:
                        row_data = []
                        is_row_null = True
                        for cell in row:
                            value = cell.value
                            if type(cell.value) == type('string'):
                                value = cell.value.strip()
                            if 'date' in str(cell.type).lower():
                                value = str(cell.value).strip()
                                if value != '':
                                    value = str(cell.value.date())
                            row_data.append(value)
                            if is_row_null and value != None and value !='':
                                is_row_null = False
                        if not is_row_null:
                            data.append(row_data)

                if '.xlsx' in str(request.FILES['file']):
                    file_type = 'xlsx'
                    book = openpyxl.load_workbook(request.FILES['file'], read_only = True)
                    sheet = book.active
                    row_count = sheet.max_row
                    column_count = sheet.max_column

                    for row in range(1, row_count+1):
                        row_data = []
                        is_row_null = True
                        for col in range(1, column_count+1):
                            value = sheet.cell(row=row, column=col).value
                            if value:
                                if sheet.cell(row=row, column=col).number_format == '0.00%':
                                    value = value * 100
                                row_data.append(str(value))
                            else:
                                row_data.append('')
                            if is_row_null and value != None and value !='':
                                is_row_null = False
                        if not is_row_null:
                            data.append(row_data)

                if len(data) > 5001:
                    return HttpResponse(AppResponse.msg(0, 'Maximum 5000 records are allowed in import.'))

                return HttpResponse(AppResponse.get({'code' : 1, 'data': data, 'mapping_columns': mapping_columns,'delimiter':None,'file_type':file_type, 'fields': fields, 'dropdown_fields':dropdown_fields, 'model_name':model_name}), content_type='json')
    except Exception as e:
        manager.create_from_exception(e)
        logging.exception('Something went wrong.')
        return HttpResponse(AppResponse.msg(0, str(e)), content_type='json')


def export_sample_product(request, group_id):
    try:
        f = BytesIO()
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Products')
        row_number = 0
        headers = [
                    'Product name', 
                    'Article Number',
                    'Internal category', 
                    'Website category', 
                    'Saleable', 
                    'Purchasable', 
                    'Type', 
                    'Sale price', 
                    'Cost price', 
                    'Traceability',
                    'Buy', 
                    'Make to order', 
                    'Make to stock', 
                    'Weight net', 
                    'Weight gross', 
                    'Desc sale', 
                    'Desc purchase', 
                    'Tax code',
                    'Tax (%)',
                    'Supplier name', 
                    'Supplier product name', 
                    'Supplier product code', 
                    'Supplier priority', 
                    'Supplier remarks',
                    'Product group'
                ]

        group = ProductGroup_Form.objects.filter(product_group_id = group_id).first()
        if group:
            product_attrs = Template_Form_Field.objects.filter(template_form = group.template_form)
            for product_attr in product_attrs:
                name = (product_attr.name).replace(" ","_")
                attribute = group.name+"_"+ name
                headers.append((attribute).lower()) 

        column_count = 0
        for header in headers:
            ws.write(row_number, column_count, header)
            column_count = column_count + 1
        row_number = row_number + 1
        wb.save(f)
        return Util.xls_to_response(wb,'products.xls')
    except Exception as e:
        manager.create_from_exception(e)
        logging.exception('Something went wrong.')
        return HttpResponse(AppResponse.msg(0, str(e)), content_type='json')